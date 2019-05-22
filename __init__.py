import openpyxl

wb = openpyxl.load_workbook("posterlognamefileblah.xlsx", "r")
sheet = wb.active
rows = sheet.max_row
col = sheet.max_column

rows = rows+1

# writing previously captured data to excel

